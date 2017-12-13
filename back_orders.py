import openpyxl
import csv
import time


#doc_file = input("Warehouse records filename:")
#files = ["CSV 11-27-2017.xlsx", "CSV 11-24-2017.xlsx"]
file = "Leanoutput.xlsx"

col_names = ["Client", "Order Id", "Tracking", "Date of Shipment",
"Product Sku #", "Quantity", "Price"] 

wb = openpyxl.load_workbook(file)
open_sheet = wb['Commercehub Open']
order_sheet = wb['Lean Supply']


open_orders = list()

# Make list of every open order
start_row = 2
end_row = open_sheet.max_row

for row in range(start_row, end_row + 1):
	
	client = str(open_sheet["BD" + str(row)].value)
	order_id = str(open_sheet["BK" + str(row)].value)
	#tracking = int(sheet["BM" + str(row)].value)
	#shipment_date = 0
	sku = str(open_sheet["CR" + str(row)].value)
	qty = int(open_sheet["BM" + str(row)].value)
	price = float(open_sheet["CN" + str(row)].value)

	open_orders.append([client, order_id, sku, qty, price])
	print(order_id, 'added')

print('All open orders compiled.')
time.sleep(3)

start_row = 2
lean_end_row = order_sheet.max_row

tracking_found = 0
tracking_not_found = 0

for order in open_orders:
	open_po = order[1]
	print(open_po)
	found = False
	for row in range(start_row, lean_end_row):
		lean_po = str(order_sheet["C" + str(row)].value)
		if open_po == lean_po:
			print('FOUND', open_po)
			found = True
			tracking = order_sheet["P" + str(row)].value
			tracking_found += 1
			if tracking == None:
				tracking = 'NO TRACKING #'
				tracking_found -= 1
				tracking_not_found += 1
			shipment_date = str(order_sheet["E" + str(row)].value)[:10]
			order.insert(2, tracking)
			order.insert(3, shipment_date)
			break
	if found == False:
		order.insert(2, '')
		order.insert(3, '')
		tracking_not_found += 1

print(open_orders)

output_wb = openpyxl.Workbook()
output_sheet = output_wb.active
col_width = 20
format_sheet = wb['Format']
for col in range(1, format_sheet.max_column + 1):
	col_letter = openpyxl.cell.cell.get_column_letter(col)
	output_sheet[col_letter + str(1)] = format_sheet[col_letter + str(1)].value
	output_sheet.column_dimensions[col_letter].width = col_width

start_row = 1
end_row = len(open_orders) + 1
for row in range(len(open_orders)):
	for col in range(1, format_sheet.max_column + 1):
		col_letter = openpyxl.cell.cell.get_column_letter(col)
		output_sheet[col_letter + str(row + 1)] = open_orders[row][col - 1]
	print('added', row)

output_wb.save('backorders.xlsx')


# with open("backorders.xlsx", "w+", newline='') as csvFile:

# 	for order in open_orders:
# 		writer.writerow(order)

print('Done.')
print('Orders with tracking numbers found:', str(tracking_found))
print('Orders with tracking numbers NOT found', str(tracking_found))